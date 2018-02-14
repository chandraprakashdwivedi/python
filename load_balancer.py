import paramiko
import re
from openpyxl import load_workbook
from openpyxl.compat import range

wb = load_workbook(filename ='157.xlsx')
sheet_ranges = wb['157']
import time
device="10.89.216.157"
conn_pre = paramiko.SSHClient()
conn_pre.set_missing_host_key_policy(paramiko.AutoAddPolicy())
conn_pre.connect(device, username="array", password="acer")
time.sleep(1)
conn = conn_pre.invoke_shell()
time.sleep(1)
output = []
commands = ["enable\n", "\n", "sh statistics cpu\n","sh llb link st\n","sh llb link ba\n"]

for command in commands:
	conn.send(command + "\r")
	time.sleep(1)
	output.append(conn.recv(100000))
str1=''.join(str(x) for x in output)
a="sh statistics cpu\\r\\n\\rCPU Utilization "
b=" %\\"
CPUUtil=str1[(str1.index(a)+len(a)):str1.index(b)]
print(int(CPUUtil))
c="Down Time"
d="Health Checkers"
Uptime=str1[(str1.index(c)+len(c)):str1.index(d)]
l=Uptime.split()
Status=l[4]
Time=l[5]
print(Status)
print(Time)
ws = wb.active
ws['A1']=str(CPUUtil)
ws['B1']=str(Status)
ws['C1']=str(Time)


e="Inbound   Avg:"
f="Inbound  Peak:"
g="Outbound   Avg:"
h="Outbound  Peak:"

a1=str1[(str1.index(e)+len(e)):str1.index(f)]
a2=str1[(str1.index(g)+len(g)):str1.index(h)]
s1=a1.replace("\\r\\n", "")
t1=a2.replace("\\r\\n", "")

p1=s1.split()
q1=t1.split()
p11=p1[1]
p12=p1[2]

q11=q1[1]
q12=q1[2]
print(p11,p12)
print(q11,q12)
str2=str1[str1.index('Bandwidth statistics:'):]
str2=str2[str2.index('PRI-LINK'):]

b1=str2[(str2.index(e)+len(e)):str1.index(f)]
b2=str2[(str2.index(g)+len(g)):str1.index(h)]
s2=b1.replace("\\r\\n", "")
t2=b2.replace("\\r\\n", "")

p2=s2.split()
q2=t2.split()
p21=p2[1]
p22=p2[2]


q21=q2[1]
q22=q2[2]

print(p21,p22)
print(q21,q22)

ws['A3']=str(p21)
ws['B3']=str(p22)
ws['A4']=str(q21)
ws['B4']=str(q22)
ws['A7']=str(p11)
ws['B7']=str(p12)
ws['A8']=str(q11)
ws['B8']=str(q12)

device="10.89.216.158"
conn_pre = paramiko.SSHClient()
conn_pre.set_missing_host_key_policy(paramiko.AutoAddPolicy())
conn_pre.connect(device, username="array", password="acer")
time.sleep(1)
conn = conn_pre.invoke_shell()
time.sleep(1)
output = []
commands = ["enable\n", "\n", "sh statistics cpu\n","sh llb link st\n","sh llb link ba\n"]

for command in commands:
	conn.send(command + "\r")
	time.sleep(1)
	output.append(conn.recv(100000))
str1=''.join(str(x) for x in output)
a="sh statistics cpu\\r\\n\\rCPU Utilization "
b=" %\\"
CPUUtil=str1[(str1.index(a)+len(a)):str1.index(b)]
print(int(CPUUtil))
c="Down Time"
d="Health Checkers"
Uptime=str1[(str1.index(c)+len(c)):str1.index(d)]
l=Uptime.split()
Status=l[4]
Time=l[5]
print(Status)
print(Time)
ws = wb.active
ws['A2']=str(CPUUtil)
ws['B2']=str(Status)
ws['C2']=str(Time)


e="Inbound   Avg:"
f="Inbound  Peak:"
g="Outbound   Avg:"
h="Outbound  Peak:"

a1=str1[(str1.index(e)+len(e)):str1.index(f)]
a2=str1[(str1.index(g)+len(g)):str1.index(h)]
s1=a1.replace("\\r\\n", "")
s1=s1.replace(" (bps)","(bps)")
t1=a2.replace("\\r\\n", "")
t1=t1.replace(" (bps)","(bps)")

p1=s1.split()
q1=t1.split()
p11=p1[1]
p12=p1[2]

q11=q1[1]
q12=q1[2]
print(p11,p12)
print(q11,q12)
str2=str1[str1.index('Bandwidth statistics:'):]
str2=str2[str2.index('PRI-LINK'):]

b1=str2[(str2.index(e)+len(e)):str1.index(f)]
b2=str2[(str2.index(g)+len(g)):str1.index(h)]
s2=b1.replace("\\r\\n", "")
s2=s2.replace(" (bps)","(bps)")
t2=b2.replace("\\r\\n", "")
t2=t2.replace(" (bps)","(bps)")
p2=s2.split()
q2=t2.split()
p21=p2[1]
p22=p2[2]


q21=q2[1]
q22=q2[2]

print(p21,p22)
print(q21,q22)

ws['A5']=str(p21)
ws['B5']=str(p22)
ws['A6']=str(q21)
ws['B6']=str(q22)
ws['A9']=str(p11)
ws['B9']=str(p12)
ws['A10']=str(q11)
ws['B10']=str(q12)
wb.save('157.xlsx')
