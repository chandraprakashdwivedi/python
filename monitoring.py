import paramiko
from openpyxl import load_workbook
from openpyxl.compat import range

wb = load_workbook(filename ='Monitoring.xlsx')
sheet_ranges = wb['Sheet']
li = ['10.88.238.6', '10.88.238.7', '10.88.238.8', '10.88.238.9', '10.88.238.71', '10.88.238.44', '10.88.238.45', '10.88.238.18', '10.88.238.19']
xlx={"10.88.238.6":["B","c"],"10.88.238.7":["G","H"],"10.88.238.8":["L","M"],"10.88.238.9":["N","O"],"10.88.238.71":["P","Q"],
    "10.88.238.44":["R","S"],"10.88.238.45":["T","U"],"10.88.238.18":["V","W"],"10.88.238.19":["AA","AB"]}
time=input("Enter Time: ")
TimeD={"2am":5,"4am":6,"6am":7,"8am":8,"10am":9,"12pm":10,"2pm":11,"4pm":12,"6pm":13,"8pm":14,"10pm":15,"12am":16}
row=TimeD[time]
ws = wb.active


def dataadd(ip,data):
    datalist=xlx[ip]

    for i in range(len(data)):
        ws[datalist[i]+str(row)]=data[i]
    

for i in range(0,len(li)):
    data=[]
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    ssh.connect(li[i], username='test', password='abc#123')
    stdin,stdout,stderr = ssh.exec_command('sh /tmp/mem; top -b -d1 -n4|grep -i "Cpu(s)" ')
    a=stdout.readlines()
    output=''.join(a)
    FindMem1,FindMem2="Percentage=","Cpu(s)"
    Memory=(output[(output.index(FindMem1)+len(FindMem1)):output.index(FindMem2)]).strip()
    FindCpu1,FindCpu2="Cpu(s):","us,"
    output=output.replace(FindCpu1," ", 3)
    output=output.replace(FindCpu2," ", 3)
    cpu=(output[(output.index(FindCpu1)+len(FindCpu1)):output.index(FindCpu2)]).strip()
    if cpu=="0.0%":
        cpu="0.10%"
    
    data.append(cpu)
    data.append(Memory)
    dataadd(li[i],data)
    
wb.save('Monitoring.xlsx')
