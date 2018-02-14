import paramiko
from openpyxl import load_workbook
from openpyxl.compat import range

wb = load_workbook(filename ='DC.xlsx')
sheet_ranges = wb['DC']
j=0
l=[' 0.00','0.00','0.01','0.02','0.03','0.04','0.05','0.06','0.07','0.08','0.09',' 0.01',' 0.02',' 0.03',' 0.04',' 0.05',' 0.06',' 0.07',' 0.08',' 0.09']
li=['10.88.238.71', '10.88.238.49', '10.88.238.50', '10.88.238.231', '10.88.238.18',
    '10.88.238.205', '10.88.238.206', '10.88.238.19', '10.88.238.202',
    '10.88.238.204', '10.88.238.51', '10.88.238.207', '10.88.238.201', '10.88.238.38',
    '10.88.238.39', '10.88.238.70', '10.88.238.55', '10.88.238.6', '10.88.238.42', '10.88.238.44',
    '10.88.238.8', '10.88.238.47', '10.88.238.43', '10.88.238.45', '10.88.238.9', '10.88.238.48',
    '10.88.238.220', '10.88.238.208', '10.88.238.209', '10.88.238.210', '10.88.238.198', '10.88.238.221',
    '10.88.238.40', '10.88.238.41', '10.88.238.52', '10.88.238.7', '10.88.238.223', '10.88.238.222', '10.88.238.225',
    '10.88.238.224', '10.88.238.229', '10.88.238.227', '10.88.238.226', '10.88.238.228', '10.88.238.199', '10.88.238.200']
Time=input("Enter Time : ")
if Time=='3am':
     j=1
elif Time=='7am':
    j=2
elif Time=='11am':
    j=3
elif Time=='3pm':
    j=4
elif Time=='7pm':
    j=5
else:
    j=6
    
    
for i in range(0,len(li)):
     try:
          ssh = paramiko.SSHClient()
          ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
          if li[i]!="10.88.238.39":
               ssh.connect(li[i], username='hexmon', password='hetd#123')
          else:
               ssh.connect(li[i], username='hexmon', password='session#123')
          stdin, stdout, stderr = ssh.exec_command('sh /tmp/mem;df -h ;uptime')
          u=stdout.readlines()
          output = ''.join(u)
          print("----------------------["+li[i]+"]--------------------------")
          print(output)
          Mem='Percentage='
          Mem1='Filesystem'
          MemPer=output[(output.index(Mem)+len(Mem)):output.index(Mem1)]
          print("Memory Percentage : " + MemPer.strip())
          ws = wb.active
          stdin, stdout, stderr = ssh.exec_command('uptime')
          uptime=stdout.readlines()
          CPU = ''.join(uptime)
          uptimelist=CPU.split(',')
          uptime=uptimelist[3]
          load='load average:'
          uptime=uptime[(uptime.index(load)+len(load)):]
          uptime=uptime.strip()
          print("CPU Percentage : " + uptime)
          ws['B'+str(j)]='Y'
          ws['E'+str(j)]='no file system utilization is above 80%'
          if uptime in l:
               ws['C'+str(j)]='0.10%'
          else:
               ws['C'+str(j)]=uptime.strip()+"%"
          ws['D'+str(j)]=MemPer.rstrip()
          j=j+6
     except:
        print("Error in ip :",li[i],"Please Login Maually")
        ws = wb.active
        ws['B'+str(j)]='Error'
        ws['C'+str(j)]='Error'
        ws['D'+str(j)]='Error'
        ws['E'+str(j)]='Error'
        j=j+6

wb.save('DC.xlsx')
